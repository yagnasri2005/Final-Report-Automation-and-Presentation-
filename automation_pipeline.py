"""
Day 28-29: Automated Data Pipeline (Simple)
--------------------------------------------
Loads raw sales data, cleans it, saves processed data, calculates key KPIs,
and exports everything to a formatted Excel workbook.

Usage:
    python automation_pipeline.py

Designed to run unattended (e.g. via GitHub Actions or Windows Task Scheduler) —
see .github/workflows/pipeline.yml or the Task Scheduler instructions in README.md.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("pipeline.log")]
)
log = logging.getLogger(__name__)

RAW_DATA_PATH = "cleaned_superstore.csv"   # source ("raw") data for this pipeline run
PROCESSED_DATA_PATH = "processed_data.csv"
EXCEL_OUTPUT_PATH = "sales_kpi_report.xlsx"


def load_raw_data(path: str) -> pd.DataFrame:
    log.info(f"Loading raw data from '{path}'...")
    df = pd.read_csv(path, parse_dates=["order_date", "ship_date"])
    log.info(f"Loaded {len(df):,} rows.")
    return df


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    log.info("Cleaning data...")
    before = len(df)

    df = df.drop_duplicates()
    df = df.dropna(subset=["order_id", "sales", "profit"])
    df = df[df["sales"] > 0]
    df["discount"] = df["discount"].clip(0, 1)

    log.info(f"Cleaning complete: {before:,} -> {len(df):,} rows "
              f"({before - len(df)} removed).")
    return df


def save_processed_data(df: pd.DataFrame, path: str) -> None:
    df.to_csv(path, index=False)
    log.info(f"Saved processed data to '{path}'.")


def calculate_kpis(df: pd.DataFrame) -> dict:
    log.info("Calculating KPIs...")
    kpis = {
        "Total Sales": round(df["sales"].sum(), 2),
        "Total Profit": round(df["profit"].sum(), 2),
        "Total Orders": df["order_id"].nunique(),
        "Total Customers": df["customer_id"].nunique(),
        "Average Order Value": round(df["sales"].mean(), 2),
        "Overall Profit Margin %": round(df["profit"].sum() / df["sales"].sum() * 100, 2),
        "Top Category": df.groupby("category")["sales"].sum().idxmax(),
        "Top Region": df.groupby("region")["profit"].sum().idxmax(),
    }
    return kpis


def category_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    return (df.groupby("category")
              .agg(total_sales=("sales", "sum"), total_profit=("profit", "sum"),
                   orders=("order_id", "nunique"))
              .round(2).reset_index()
              .sort_values("total_sales", ascending=False))


def region_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    return (df.groupby("region")
              .agg(total_sales=("sales", "sum"), total_profit=("profit", "sum"),
                   orders=("order_id", "nunique"))
              .round(2).reset_index()
              .sort_values("total_profit", ascending=False))


def monthly_trend(df: pd.DataFrame) -> pd.DataFrame:
    ts = df.set_index("order_date")["sales"].resample("ME").sum().round(2)
    return ts.reset_index().rename(columns={"order_date": "month", "sales": "total_sales"})


def export_to_excel(kpis: dict, cat_df: pd.DataFrame, region_df: pd.DataFrame,
                     trend_df: pd.DataFrame, path: str) -> None:
    log.info(f"Exporting KPI report to '{path}'...")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        kpi_df = pd.DataFrame(list(kpis.items()), columns=["KPI", "Value"])
        kpi_df.to_excel(writer, sheet_name="KPI Summary", index=False)
        cat_df.to_excel(writer, sheet_name="By Category", index=False)
        region_df.to_excel(writer, sheet_name="By Region", index=False)
        trend_df.to_excel(writer, sheet_name="Monthly Trend", index=False)

    # Basic formatting pass
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(path)
    header_fill = PatternFill(start_color="2EC4B6", end_color="2EC4B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col if c.value is not None)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4
    wb.save(path)
    log.info("Excel export complete.")


def run_pipeline():
    log.info("="*60)
    log.info(f"PIPELINE RUN STARTED — {datetime.now().isoformat()}")
    log.info("="*60)

    df = load_raw_data(RAW_DATA_PATH)
    df = clean_data(df)
    save_processed_data(df, PROCESSED_DATA_PATH)

    kpis = calculate_kpis(df)
    cat_df = category_breakdown(df)
    region_df = region_breakdown(df)
    trend_df = monthly_trend(df)

    export_to_excel(kpis, cat_df, region_df, trend_df, EXCEL_OUTPUT_PATH)

    log.info("KPI SUMMARY:")
    for k, v in kpis.items():
        log.info(f"  {k}: {v}")

    log.info("="*60)
    log.info("PIPELINE RUN COMPLETE")
    log.info("="*60)


if __name__ == "__main__":
    run_pipeline()
